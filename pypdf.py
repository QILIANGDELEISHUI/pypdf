import re
import fitz  # PyMuPDF的导入名称是fitz
import os
import sys
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Border, Side, Font, Alignment, numbers
from openpyxl.utils import get_column_letter


def extraction_issue_date(text):
    """
    提取发票日期

    :param text: 文本内容
    :return: 发票日期
    """
    # 改进的正则表达式，允许年份、月份、日期与中文单位之间存在空格
    pattern = r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"
    match = re.search(pattern, text)

    if match:
        # 重新组合日期，去除空格
        year, month, day = match.groups()
        return f"{year}年{month}月{day}日"

    return None


def extraction_invoice_number(text):
    """
    提取发票号码

    :param text: 文本内容
    :return: 发票号码
    """
    # 1. 尝试匹配全电发票（20位数字）
    pattern_electronic = r"(?<!\d)\d{20}(?!\d)"
    match = re.search(pattern_electronic, text)
    if match:
        return match.group(0)

    # 2. 尝试匹配普通发票/专票（8位数字）- 改进版
    # 明确前缀和数字之间的分隔符，如冒号、空格等
    pattern_normal = r"(?:发票号码|NO\.?)\s*[:：\s]*(\d{8})"
    match = re.search(pattern_normal, text, re.IGNORECASE)
    if match:
        return match.group(1)  # 只返回括号内捕获的数字部分

    # 3. 直接匹配8位数字（无明确前缀时）
    pattern_standalone = r"(?<!\d)\d{8}(?!\d)"
    match = re.search(pattern_standalone, text)
    if match:
        # 增加上下文验证：检查前后是否有"发票"相关词汇
        context = text[max(0, match.start() - 20) : min(len(text), match.end() + 20)]
        if re.search(r"发票|invoice|NO\.", context, re.IGNORECASE):
            return match.group(0)

    return None


def extraction_project_name(text):
    """
    提取项目名称

    提取规则：
    1. 必须以*开头
    2. 当前行字符数量（宽度）大于等于22时，继续提取后续行
    3. 后续行不能以*或￥开头
    4. 若当前行字符数量不足22且不以*或￥开头，则结束提取
    """
    lines = text.split("\n")

    def calculate_width(line):
        """计算文本行的宽度（中文字符宽度为2，英文字符宽度为1）"""
        return sum(2 if ord(char) > 127 else 1 for char in line)

    project_name_parts = []
    in_project_section = False

    for line in lines:
        line = line.strip()
        line_width = calculate_width(line)

        # 如果不在项目部分，寻找起始行（以*开头）
        if not in_project_section:
            if line.startswith("*"):
                in_project_section = True
                project_name_parts.append(line)

        # 如果已在项目部分，处理后续行
        else:
            # 若当前行以*或￥开头，结束提取
            if line.startswith(("*", "¥")):
                break

            # 若当前行宽度不足22，结束提取
            if line_width < 22:
                break

            # 符合条件的行，添加到项目名称
            project_name_parts.append(line)

    return " ".join(project_name_parts)


def extraction_amount(text):
    """
    提取金额

    :param text: 文本内容
    :return: 金额
    """
    # 同时匹配 ¥ 和 ￥ 符号，忽略符号后的空格
    total_amount_matches = re.findall(r"[¥￥]\s*(\d+\.\d{2})", text)

    # 处理匹配结果
    if not total_amount_matches:
        return None  # 没有匹配到任何金额

    # 如果只有一个匹配，直接返回
    if len(total_amount_matches) == 1:
        try:
            return float(total_amount_matches[0])
        except ValueError:
            return None

    # 如果有多个匹配，尝试返回第三个（索引为2）
    try:
        return float(total_amount_matches[2])
    except (IndexError, ValueError):
        return None


def extract_invoice_data(pdf_path):
    """
    从 PDF 文件中提取发票信息

    :param pdf_path: PDF 文件路径
    :return: 包含发票信息的字典
    """
    try:
        # 打开 PDF 文件（自动处理文件路径和异常）
        with fitz.open(pdf_path) as doc:
            # 检查 PDF 是否有至少一页
            if doc.page_count < 1:
                raise ValueError("PDF 无内容或页数不足")

            # 获取第一页（索引从 0 开始）
            page = doc[0]  # 替代 load_page(0)
            text = page.get_textpage().extractText()

            # print(text)

            # 提取开票日期
            issue_date = extraction_issue_date(text)
            # 提取发票号码
            invoice_number = extraction_invoice_number(text)
            # 提取项目名称
            project_name = extraction_project_name(text)
            # 提取价税合计（小写）
            total_amount = extraction_amount(text)

            # 将提取的信息存入字典
            data = {
                "发票号码": invoice_number,
                "开票日期": issue_date,
                "报销项目": project_name,
                "价税合计": total_amount,
            }

            return data

    except FileNotFoundError:
        print(f"错误：文件 {pdf_path} 未找到")
        raise  # 重新抛出异常，通知调用者
    except fitz.FileDataError as e:
        print(f"PyMuPDF 错误：{e}，可能文件损坏或非 PDF 格式")
        raise  # 重新抛出异常
    except Exception as e:
        print(f"其他错误：{e}")
        raise  # 重新抛出异常


def traverse_pdf_files(document_dir):
    """
    遍历指定目录下的所有 PDF 文件

    :param document_dir: 文件目录
    :return: 包含所有 PDF 文件名列表
    """
    pdf_data = []
    for filename in os.listdir(document_dir):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(document_dir, filename)
            try:
                # 尝试提取发票数据
                invoice_data = extract_invoice_data(pdf_path)
                pdf_data.append((filename, invoice_data))
            except FileNotFoundError as e:
                # 处理文件不存在异常
                print(f"错误：文件 {pdf_path} 不存在")
                raise
            except PermissionError as e:
                # 处理权限不足异常
                print(f"错误：没有权限访问文件 {pdf_path}")
                raise
            except Exception as e:
                # 处理其他未知异常
                print(f"错误：处理文件 {pdf_path} 时发生异常: {str(e)}")
                raise
    return pdf_data


def write_to_excel(pdf_data, excel_path):
    """
    将提取的发票信息写入 Excel 文件

    :param pdf_data: 包含发票信息的列表
    :param excel_path: Excel 文件路径
    """
    # openpyxl 使用的是“写入时构建”的方法，
    # 即在你调用 ws.append() 方法时，数据会被直接写入到内存中的一个数据结构中，而不是整个Excel文件。
    # 如果后期数据量过大，可以考虑使用 openpyxl 的优化方法。如 write_only 模式。
    wb = Workbook()
    if wb is None:
        raise RuntimeError("无法创建 Excel 工作簿")

    # 获取活动工作表（默认存在）
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法获取工作表")
    header = ["文件名", "发票号码", "开票日期", "项目名称", "价税合计（小写）"]
    ws.append(header)

    # 设置表头样式：加粗且居中
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    for filename, invoice_data in pdf_data:
        row = [filename] + list(invoice_data.values())
        ws.append(row)

    # 适应列宽
    for column in ws.columns:
        max_length = 0
        # 确保column[0].column不是None
        if column[0].column is not None:
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    cell_value = str(cell.value)
                    # 计算字符宽度，中文字符宽度乘以 2
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    if length > max_length:
                        max_length = length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    # 添加框线
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    wb.save(excel_path)


def resource_path(relative_path):
    """
    获取资源文件的绝对路径..

    :param relative_path: 相对路径
    :return: 绝对路径
    """
    if getattr(sys, "frozen", False):
        # 运行在打包后的exe中
        base_path = os.path.dirname(sys.executable)
        print(f"打包后 base_path: {os.path.join(base_path, relative_path)}")
    else:
        # 运行在普通Python环境中
        base_path = os.path.dirname(os.path.abspath(__file__))
        print(f"普通环境 base_path: {os.path.join(base_path, relative_path)}")
    return os.path.join(base_path, relative_path)


def main():
    try:
        # 获取当前时间
        now = datetime.now()
        # 格式化时间字符串
        formatted_time = now.strftime("%Y年%m月%d日%H点%M分导出")
        # 构建 document 文件夹的路径 源数据文件夹
        document_dir = resource_path("document")
        if not os.path.exists(document_dir):
            raise FileNotFoundError("document 文件夹不存在，请检查路径是否正确")
        # 构建 data 文件夹的路径 结果数据文件夹
        data_dir = resource_path("data")
        # 如果 data 文件夹不存在，则创建
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        # 构建 Excel 文件路径 实时时间中不包含秒，短时间内重复运行可能会导致文件名重复
        excel_path = os.path.join(data_dir, f"{formatted_time}导出发票信息.xlsx")
        pdf_data = traverse_pdf_files(document_dir)
        write_to_excel(pdf_data, excel_path)
        print(f"发票信息已保存到 {excel_path}")
    except Exception as e:
        print(f"main发生错误: {e}")
    input("按任意键退出...")


if __name__ == "__main__":
    main()
    # resource_path("document")
    # path = "document/7.功放模块 8.2.pdf"
    # path = "document/3.高速电机 59.5.pdf"
    # extract_invoice_data(path)
